import os
from dotenv import load_dotenv
from pathlib import Path
from openpyxl import load_workbook
env_path = Path(__file__).parents[1] / '.env'
load_dotenv(env_path)
path = os.getenv("xlsx_path")
def main():
    wb = load_workbook(path)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if row[0] is not None:
            print(row)
    
    
    
if __name__ == "__main__":
    main()